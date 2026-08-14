"""Локальный реестр на месте Google Таблицы.

Нужен для одного: чтобы проект можно было запустить, имея только ключ
модели. Настройка сервисного аккаунта Google занимает полчаса, и человек,
который просто хочет посмотреть, до конца её обычно не доводит.

Формат тот же — три листа с теми же колонками, — поэтому переключение
между локальным файлом и Google Таблицей ничего не меняет в логике выше.
"""

import shutil
import threading
from pathlib import Path

from openpyxl import Workbook, load_workbook

import config

_lock = threading.Lock()


def _workbook_path():
    """Рабочая копия, чтобы эталонные демо-данные в репозитории не менялись."""
    path = Path(config.DEMO_FILE)
    if not path.exists():
        shutil.copy(config.DEMO_SOURCE, path)
    return path


def read(sheet_name, columns):
    """Список словарей с номером строки, как у Google-версии."""
    with _lock:
        wb = load_workbook(_workbook_path(), data_only=True)
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]

        rows = []
        for offset, raw in enumerate(ws.iter_rows(min_row=2, max_col=len(columns))):
            values = ["" if c.value is None else str(c.value).strip() for c in raw]
            if not any(values):
                continue
            record = dict(zip(columns, values))
            record["_row"] = offset + 2
            rows.append(record)
        return rows


def append(sheet_name, rows):
    """Дописывает строки в конец листа."""
    with _lock:
        path = _workbook_path()
        wb = load_workbook(path)
        ws = wb[sheet_name]
        for row in rows:
            ws.append(row)
        wb.save(path)


def update(sheet_name, changes):
    """changes — список (номер строки, буква колонки, значение)."""
    with _lock:
        path = _workbook_path()
        wb = load_workbook(path)
        ws = wb[sheet_name]
        for row, column, value in changes:
            ws[f"{column}{row}"] = value
        wb.save(path)


def add_event(title, date, description, time):
    """Событие календаря пишется в отдельный лист.

    Заменять Google Calendar полноценно незачем, но действие должно быть
    видимым: человек открывает файл и находит там созданную запись.
    """
    with _lock:
        path = _workbook_path()
        wb = load_workbook(path)
        if "calendar" not in wb.sheetnames:
            ws = wb.create_sheet("calendar")
            ws.append(["date", "time", "title", "description"])
        ws = wb["calendar"]
        ws.append([date, time or "весь день", title, description])
        wb.save(path)
        return {"id": f"demo-{ws.max_row}", "link": ""}


def reset():
    """Возвращает демо-данные в исходное состояние."""
    with _lock:
        path = Path(config.DEMO_FILE)
        if path.exists():
            path.unlink()
        shutil.copy(config.DEMO_SOURCE, path)
